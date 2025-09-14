"""
Exporte des données depuis MariaDB vers un fichier Excel prêt à ouvrir dans Excel (Mac ou Windows),
sans ODBC, en utilisant SQLAlchemy + pandas.

Feuilles générées :
  - "latest"     : vue v_portfolio_latest (photo courante)
  - "timeseries" : total du portefeuille dans le temps (ts, total_eur, investi_eur)
  - "history"    : historique détaillé (optionnel, peut être lourd)

Sortie : reports/portfolio_latest.xlsx

Dépendances Python : pandas, SQLAlchemy, PyMySQL, python-dotenv (facultatif)
Variables d'environnement DB_* lues par src.db.get_engine() :
  DB_HOST, DB_PORT, DB_USER, DB_PASSWORD, DB_NAME
---
Changelog:
- 2025-09-12 19:40 (Europe/Paris) — [Aya] Création du script d’export Excel (`export_latest_to_excel.py`) avec 3 feuilles (latest, timeseries, history) et auto-formatage openpyxl.
"""

from pathlib import Path
import os
import pandas as pd

# On réutilise l'engine existant du projet
try:
    from src.db import get_engine
except ImportError:
    from db import get_engine  # fallback si l'import relatif n'est pas utilisé


def _autofit_openpyxl(ws):
    """Ajuste la largeur des colonnes pour la feuille openpyxl (best-effort)."""
    from openpyxl.utils import get_column_letter

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            try:
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        # marge visuelle
        adjusted = min(max_len + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def export_excel(out_path: Path, include_history: bool = True) -> Path:
    eng = get_engine()

    # Requêtes principales
    sql_latest = (
        "SELECT ts, symbol, qty, investi, price_eur, valeur_actuelle, pnl_eur, pnl_pct "
        "FROM v_portfolio_latest ORDER BY symbol"
    )

    sql_timeseries = (
        "SELECT ts, SUM(valeur_actuelle) AS total_eur, SUM(investi) AS investi_eur "
        "FROM portfolio_snapshot GROUP BY ts ORDER BY ts"
    )

    sql_history = (
        "SELECT ts, symbol, qty, investi, price_eur, valeur_actuelle, pnl_eur, pnl_pct "
        "FROM portfolio_snapshot ORDER BY ts, symbol"
    )

    print("[DB] Lecture des données…")
    df_latest = pd.read_sql(sql_latest, eng)
    df_timeseries = pd.read_sql(sql_timeseries, eng)
    df_history = pd.read_sql(sql_history, eng) if include_history else None

    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"[XLSX] Écriture → {out_path}")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_latest.to_excel(writer, sheet_name="latest", index=False)
        df_timeseries.to_excel(writer, sheet_name="timeseries", index=False)
        if df_history is not None:
            df_history.to_excel(writer, sheet_name="history", index=False)

        # Mise en forme simple : auto-fit des colonnes + filtres auto
        wb = writer.book
        for name in ("latest", "timeseries") + (
            ("history",) if df_history is not None else tuple()
        ):
            ws = wb[name]
            # Filtre automatique sur la première ligne si données présentes
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            _autofit_openpyxl(ws)

    print("[OK] Fichier Excel généré.")
    return out_path


def main():
    # Destination par défaut
    out = Path("reports/portfolio_latest.xlsx")

    # Permettre de désactiver l'onglet history via variable d'env (pour gros jeux de données)
    include_history = os.getenv("PICSOU_INCLUDE_HISTORY", "1") not in {
        "0",
        "false",
        "False",
    }

    export_excel(out, include_history=include_history)
    print("Ouvre ce fichier dans Excel et fais tes graphiques/segments dessus ✨")


if __name__ == "__main__":
    main()
